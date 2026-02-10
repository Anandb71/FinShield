"""
FinShield - Excel Normalization Worker (The "Repair Shop")

Handles broken Excel inputs before extraction:
- Detects "Unnamed" columns
- Fixes headers (infers Date, Amount, Balance from data)
- Cleans garbage text (OCR artifacts like "unrings ICEASE")
- Handles multi-section files (Account 3 summary + transactions)
- Normalizes date formats across merged files
- Detects and extracts opening/closing balances from summary rows
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from dateutil import parser as dateparser


@dataclass
class NormalizedStatement:
    """Result of normalizing an Excel bank statement."""
    opening_balance: Optional[float] = None
    closing_balance: Optional[float] = None
    account_number: Optional[str] = None
    account_holder: Optional[str] = None
    period_from: Optional[str] = None
    period_to: Optional[str] = None
    transactions: List[Dict[str, Any]] = field(default_factory=list)
    repair_log: List[str] = field(default_factory=list)
    raw_headers: List[str] = field(default_factory=list)
    detected_anomalies: List[Dict[str, Any]] = field(default_factory=list)
    # Metadata integrity: stores header vs calculated discrepancy for fraud detection
    metadata_discrepancy: Optional[Dict[str, Any]] = None


# Common header aliases for bank statements
HEADER_ALIASES = {
    "date": ["tran date", "txn date", "transaction date", "date", "value date", "posting date"],
    "description": ["particulars", "description", "narration", "transaction details", "details", "memo"],
    "debit": ["debit", "withdrawal", "withdrawals", "dr", "debit amount"],
    "credit": ["credit", "deposit", "deposits", "cr", "credit amount"],
    "balance": ["balance", "closing balance", "balanc e", "running balance", "balance after"],
    "cheque": ["chq no", "cheque no", "chq no.", "ref no./cheque no.", "ref no", "reference"],
}


def _safe_float(value: Any) -> Optional[float]:
    """Parse a value to float, handling commas and currency symbols."""
    if value is None or value == "" or value == "None":
        return None
    s = str(value).strip()
    # Remove currency symbols and commas
    s = re.sub(r"[₹$€£,\s]", "", s)
    # Handle parenthesized negatives like (1000)
    m = re.match(r"^\(([\d.]+)\)$", s)
    if m:
        return -float(m.group(1))
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _is_date_value(value: Any) -> bool:
    """Check if a value looks like a date."""
    if value is None:
        return False
    if isinstance(value, datetime):
        return True
    s = str(value).strip()
    if not s or len(s) < 6:
        return False
    # Common date patterns
    patterns = [
        r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}",
        r"\d{1,2}\s+[A-Za-z]{3}\s+\d{2,4}",
        r"\d{4}-\d{2}-\d{2}",
        r"[A-Za-z]{3}.*\d{4}",
    ]
    return any(re.search(p, s) for p in patterns)


def _is_garbage_text(text: str) -> bool:
    """Detect OCR garbage / merge artifact text."""
    if not text:
        return False
    lower = text.lower().strip()
    # Known garbage patterns from dataset analysis
    garbage_patterns = [
        r"unrings\s+icease",
        r"pherate.*vumar",
        r"0511\s*nn",
        r"\$\d+",  # Dollar amounts in INR context
    ]
    if any(re.search(p, lower) for p in garbage_patterns):
        return True
    # Very high ratio of non-alphanumeric characters
    alnum = sum(1 for c in text if c.isalnum() or c.isspace())
    if len(text) > 5 and alnum / len(text) < 0.3:
        return True
    return False


def _detect_header_row(ws, max_scan: int = 30) -> Tuple[int, Dict[str, int]]:
    """Find the header row and map columns to semantic names."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        cells = [str(ws.cell(row=row_idx, column=c).value or "").strip().lower()
                 for c in range(1, ws.max_column + 1)]
        mapping: Dict[str, int] = {}
        for semantic, aliases in HEADER_ALIASES.items():
            for col_idx, cell_text in enumerate(cells):
                if cell_text in aliases or any(a in cell_text for a in aliases):
                    if semantic not in mapping:
                        mapping[semantic] = col_idx + 1  # 1-based
        # Need at least date + (debit or credit or description) to be a header row
        if "date" in mapping and ("debit" in mapping or "credit" in mapping or "description" in mapping):
            return row_idx, mapping
    return -1, {}


def _extract_summary_balances(ws, max_scan: int = 25) -> Tuple[Optional[float], Optional[float]]:
    """Look for OPENING BALANCE and CLOSING BALANCE in summary sections."""
    opening = None
    closing = None
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val).strip().upper()
            if "OPENING BALANCE" in text or "OPENING BAL" in text:
                # Look for numeric value in same row or next columns
                for c in range(col_idx + 1, min(ws.max_column + 1, col_idx + 5)):
                    v = _safe_float(ws.cell(row=row_idx, column=c).value)
                    if v is not None:
                        opening = v
                        break
                # Also check next cell in same row if merged
                v = _safe_float(ws.cell(row=row_idx, column=2).value)
                if opening is None and v is not None:
                    opening = v
            if "CLOSING BALANCE" in text or "CLOSING BAL" in text:
                for c in range(col_idx + 1, min(ws.max_column + 1, col_idx + 5)):
                    v = _safe_float(ws.cell(row=row_idx, column=c).value)
                    if v is not None:
                        closing = v
                        break
                v = _safe_float(ws.cell(row=row_idx, column=2).value)
                if closing is None and v is not None:
                    closing = v
    return opening, closing


def _classify_transaction(desc: str, amount: float) -> Tuple[str, str]:
    """Classify transaction into category and merchant_normalized."""
    lower = (desc or "").lower()
    category = "Other"
    merchant = ""

    if "salary" in lower or "bulk posting" in lower:
        category = "Salary"
    elif "upi" in lower:
        category = "UPI Payment"
        # Extract counterparty from UPI string
        m = re.search(r"upi/p2[am]/\d+/([^/]+)", lower)
        if m:
            merchant = m.group(1).strip().title()
    elif "neft" in lower:
        category = "NEFT Transfer"
        m = re.search(r"neft/[^/]+/([^/]+)", lower)
        if m:
            merchant = m.group(1).strip().title()
    elif "imps" in lower:
        category = "IMPS Transfer"
    elif "atm" in lower or "cash" in lower:
        category = "ATM/Cash"
    elif "emi" in lower or "loan" in lower:
        category = "EMI/Loan"
    elif any(x in lower for x in ["fee", "charge", "chrg"]):
        category = "Fees & Charges"
    elif "interest" in lower:
        category = "Interest"
    elif any(x in lower for x in ["card", "pos"]):
        category = "Card Payment"
    elif any(x in lower for x in ["insurance", "premium"]):
        category = "Insurance"
    elif any(x in lower for x in ["transfer", "trf"]):
        category = "Transfer"
    elif any(x in lower for x in ["bill", "recharge"]):
        category = "Bill Payment"
    elif amount > 0:
        category = "Income"
    elif amount < 0:
        category = "Expense"

    if not merchant and desc:
        # Fallback: use first few words cleaned up
        words = re.sub(r"[^a-zA-Z\s]", " ", desc).strip().split()[:3]
        merchant = " ".join(words).title() if words else ""

    return category, merchant


def _is_header_text(value: Any) -> bool:
    """Check if a cell value is a header alias string rather than data."""
    if value is None:
        return False
    text = str(value).strip().lower()
    if not text:
        return False
    all_aliases = set()
    for aliases in HEADER_ALIASES.values():
        for a in aliases:
            all_aliases.add(a)
    return text in all_aliases


def _try_parse_date(value: Any) -> Optional[datetime]:
    """Attempt to parse a value as a date. Returns None if not a valid date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s or len(s) < 6:
        return None
    # Reject strings that are clearly header text
    if _is_header_text(s):
        return None
    try:
        return dateparser.parse(s, dayfirst=True)
    except Exception:
        return None


def normalize_excel_statement(content: bytes, filename: str = "") -> NormalizedStatement:
    """
    Normalize a bank statement Excel file into a clean, structured format.
    
    Handles real-world messiness found in financial documents:
    - Multi-section files with summary + transaction sections
    - Mixed date formats, OCR garbage rows
    - Comma-formatted numbers, varying header names
    - Summary sections at top before transactions
    - Embedded transaction totals and metadata rows
    """
    result = NormalizedStatement()
    result.repair_log.append(f"normalizing: {filename}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        result.repair_log.append(f"failed to open workbook: {exc}")
        return result

    ws = wb.active
    if not ws:
        result.repair_log.append("no active sheet found")
        wb.close()
        return result

    result.repair_log.append(f"sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

    # Step 1: Extract summary balances from top section
    summary_opening, summary_closing = _extract_summary_balances(ws)
    if summary_opening is not None:
        result.repair_log.append(f"summary_opening_balance: {summary_opening}")
    if summary_closing is not None:
        result.repair_log.append(f"summary_closing_balance: {summary_closing}")

    # Step 2: Find the header row
    header_row, col_map = _detect_header_row(ws)
    if header_row < 0:
        result.repair_log.append("could not detect header row - attempting column inference")
        # Fallback: try common layouts
        header_row = 1
        col_map = {"date": 2, "description": 4, "debit": 5, "credit": 6, "balance": 7}

    result.repair_log.append(f"header_row: {header_row}, columns: {col_map}")
    result.raw_headers = [str(ws.cell(row=header_row, column=c).value or "") for c in range(1, ws.max_column + 1)]

    # Step 3: Find opening balance row (right after header)
    # ONLY use rows explicitly labelled "OPENING BALANCE", and read from balance column
    opening_from_row = None
    opening_row_idx = None
    for row_idx in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and "OPENING BALANCE" in str(val).upper():
                bal_col = col_map.get("balance", 7)
                opening_from_row = _safe_float(ws.cell(row=row_idx, column=bal_col).value)
                opening_row_idx = row_idx
                result.repair_log.append(f"opening_from_transactions: {opening_from_row} (row {row_idx})")
                break
        if opening_from_row is not None:
            break

    # Priority: explicit OPENING BALANCE row > summary section
    result.opening_balance = opening_from_row if opening_from_row is not None else summary_opening

    # Step 4: Parse transaction rows
    date_col = col_map.get("date", 2)
    desc_col = col_map.get("description", 4)
    debit_col = col_map.get("debit", 5)
    credit_col = col_map.get("credit", 6)
    balance_col = col_map.get("balance", 7)
    cheque_col = col_map.get("cheque")

    garbage_count = 0
    header_rows_skipped = 0
    empty_streak = 0
    date_formats_seen: Dict[str, int] = {}
    # Authoritative closing balance — set exactly once with clear priority
    closing_from_row: Optional[float] = None
    row_idx = header_row + 1

    while row_idx <= ws.max_row:
        date_val = ws.cell(row=row_idx, column=date_col).value
        desc_val = ws.cell(row=row_idx, column=desc_col).value
        debit_val = ws.cell(row=row_idx, column=debit_col).value
        credit_val = ws.cell(row=row_idx, column=credit_col).value
        balance_val = ws.cell(row=row_idx, column=balance_col).value

        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in [date_val, desc_val, debit_val, credit_val, balance_val]):
            empty_streak += 1
            if empty_streak > 10:
                break  # Likely end of data
            row_idx += 1
            continue
        empty_streak = 0

        # ── FIX #1: Skip rows where ANY cell contains header-alias text ──
        row_cells = [date_val, desc_val, debit_val, credit_val, balance_val]
        if any(_is_header_text(v) for v in row_cells):
            header_rows_skipped += 1
            result.repair_log.append(f"skipped header-text row {row_idx}")
            row_idx += 1
            continue

        # Skip summary/total rows
        desc_str = str(desc_val or "").strip().upper()
        if desc_str in ("OPENING BALANCE", "CLOSING BALANCE", "TRANSACTION TOTAL", "TOTAL", ""):
            if desc_str == "CLOSING BALANCE":
                cb = _safe_float(balance_val)
                if cb is not None and closing_from_row is None:
                    closing_from_row = cb
                    result.repair_log.append(f"closing_from_transactions: {closing_from_row}")
            row_idx += 1
            continue

        # Check for garbage text
        desc_text = str(desc_val or "")
        if _is_garbage_text(desc_text):
            garbage_count += 1
            result.detected_anomalies.append({
                "type": "garbage_text",
                "severity": "warning",
                "row": row_idx,
                "description": f"OCR garbage detected: {desc_text[:60]}",
            })
            row_idx += 1
            continue

        # ── FIX #3: Validate date before accepting the row ──
        parsed_date = _try_parse_date(date_val)
        date_str = str(date_val or "").strip()

        # If date column has text that doesn't parse as a date and there's no
        # numeric amount, this isn't a real transaction row — skip it.
        debit = _safe_float(debit_val)
        credit = _safe_float(credit_val)
        balance = _safe_float(balance_val)

        if not parsed_date and date_str and date_str != "None":
            # Date column has something, but it's not a valid date.
            # Only keep the row if it has a numeric amount (continuation row).
            if debit is None and credit is None and balance is None:
                result.repair_log.append(f"skipped non-date row {row_idx}: date='{date_str}'")
                row_idx += 1
                continue

        # Parse amount
        amount = None
        tx_type = "unknown"
        if credit is not None and credit > 0:
            amount = credit
            tx_type = "credit"
        if debit is not None and debit > 0:
            amount = -debit  # Debits are negative
            tx_type = "debit"
        # If neither, try to infer from a single amount column
        if amount is None and (debit is not None or credit is not None):
            amount = credit if credit is not None else (-debit if debit is not None else None)

        # ── FIX #1 (part 2): Skip phantom rows with no amount AND no balance ──
        # These are rows leaking from headers or empty data — not real transactions
        if amount is None and balance is None:
            result.repair_log.append(f"skipped phantom row {row_idx}: no amount or balance")
            row_idx += 1
            continue

        # Track date formats — only for rows with a valid parsed date
        if parsed_date:
            if isinstance(date_val, datetime):
                date_formats_seen["datetime_obj"] = date_formats_seen.get("datetime_obj", 0) + 1
            elif re.search(r"[A-Za-z]{3}", date_str):
                date_formats_seen["alpha"] = date_formats_seen.get("alpha", 0) + 1
            elif re.search(r"\d{4}-\d{2}", date_str):
                date_formats_seen["iso"] = date_formats_seen.get("iso", 0) + 1
            elif re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", date_str):
                date_formats_seen["numeric"] = date_formats_seen.get("numeric", 0) + 1

        # Use the parsed date for consistent ISO output
        normalized_date_str = parsed_date.strftime("%Y-%m-%d") if parsed_date else (date_str if date_str != "None" else None)

        # Classify the transaction
        category, merchant = _classify_transaction(desc_text, amount or 0)

        tx = {
            "date": normalized_date_str,
            "description": desc_text,
            "amount": amount,
            "balance": balance,
            "tx_type": tx_type,
            "category": category,
            "merchant_normalized": merchant,
            "row_index": row_idx,
        }
        if cheque_col:
            tx["cheque_no"] = str(ws.cell(row=row_idx, column=cheque_col).value or "").strip() or None

        result.transactions.append(tx)
        row_idx += 1

    # ── FIX #2: Single authoritative closing-balance computation ──
    # Priority: explicit CLOSING BALANCE row > last transaction balance > summary section
    if closing_from_row is not None:
        result.closing_balance = closing_from_row
    elif result.transactions:
        # Fallback: last transaction's running balance
        last_balance = None
        for tx in reversed(result.transactions):
            if tx.get("balance") is not None:
                last_balance = tx["balance"]
                break
        if last_balance is not None:
            result.closing_balance = last_balance
            result.repair_log.append(f"closing_inferred_from_last_row: {last_balance}")
    if result.closing_balance is None and summary_closing is not None:
        result.closing_balance = summary_closing
        result.repair_log.append(f"closing_from_summary: {summary_closing}")

    # Step 5: Metadata Integrity Check
    # Compare the header/summary closing balance against the actual last
    # transaction balance.  A massive discrepancy (>1.0 AND >50x) means the
    # header is lying — possible data tampering / summary injection.
    if result.transactions:
        last_tx_balance = None
        for tx in reversed(result.transactions):
            if tx.get("balance") is not None:
                last_tx_balance = tx["balance"]
                break

        # Also compare against the raw summary_closing extracted from the
        # header section at the top of the sheet.
        header_closing = summary_closing  # from Step 1
        calculated_closing = last_tx_balance

        if header_closing is not None and calculated_closing is not None:
            discrepancy = abs(header_closing - calculated_closing)
            if discrepancy > 1.0 and abs(calculated_closing) > 0 and abs(header_closing) > abs(calculated_closing) * 5:
                result.metadata_discrepancy = {
                    "header_closing": header_closing,
                    "calculated_closing": calculated_closing,
                    "discrepancy": discrepancy,
                    "ratio": round(abs(header_closing / calculated_closing), 2) if calculated_closing != 0 else 999999999,
                }
                result.detected_anomalies.append({
                    "type": "metadata_integrity_failure",
                    "severity": "critical",
                    "description": (
                        f"FRAUD SIGNAL: Header closing balance ({header_closing:,.2f}) "
                        f"does not match calculated row balance ({calculated_closing:,.2f}). "
                        f"Discrepancy: {discrepancy:,.2f}. "
                        "The document header is inconsistent with the transaction data — "
                        "possible summary injection or data tampering."
                    ),
                    "header_closing": header_closing,
                    "calculated_closing": calculated_closing,
                    "discrepancy": discrepancy,
                    "ratio": round(abs(header_closing / calculated_closing), 2) if calculated_closing != 0 else 999999999,
                })
                result.repair_log.append(
                    f"CRITICAL: metadata integrity failure - "
                    f"header_closing={header_closing} vs calculated={calculated_closing} "
                    f"(discrepancy={discrepancy:,.2f})"
                )
                # DO NOT override closing_balance — keep the fraudulent header
                # value so the validation layer also catches the mismatch
                # and fires its own balance-continuity error.

        # Separate check: if closing_balance was set from rows (closing_from_row)
        # but differs wildly from the last transaction balance, flag it too
        if result.closing_balance is not None and last_tx_balance is not None:
            if abs(result.closing_balance) > abs(last_tx_balance) * 50 and abs(last_tx_balance) > 0:
                if result.metadata_discrepancy is None:
                    result.metadata_discrepancy = {
                        "header_closing": result.closing_balance,
                        "calculated_closing": last_tx_balance,
                        "discrepancy": abs(result.closing_balance - last_tx_balance),
                        "ratio": round(abs(result.closing_balance / last_tx_balance), 2) if last_tx_balance != 0 else 999999999,
                    }
                    result.detected_anomalies.append({
                        "type": "metadata_integrity_failure",
                        "severity": "critical",
                        "description": (
                            f"FRAUD SIGNAL: Closing balance ({result.closing_balance:,.2f}) is wildly inconsistent "
                            f"with last transaction balance ({last_tx_balance:,.2f}). "
                            "Possible summary injection or data tampering."
                        ),
                    })
                    result.repair_log.append(
                        f"CRITICAL: summary injection detected - "
                        f"closing={result.closing_balance} vs last_row={last_tx_balance}"
                    )

    # Step 6: Check for merge artifact (mixed date formats)
    # Only flag if there are truly multiple REAL date formats (ignore single-count outliers)
    real_formats = {k: v for k, v in date_formats_seen.items() if v >= 2}
    if len(real_formats) > 1:
        result.detected_anomalies.append({
            "type": "merge_artifact",
            "severity": "warning",
            "description": f"Mixed date formats detected: {real_formats}. Possible file merge.",
        })
        result.repair_log.append(f"merge_artifact: mixed date formats {real_formats}")

    if garbage_count > 0:
        result.repair_log.append(f"garbage_rows_skipped: {garbage_count}")
    if header_rows_skipped > 0:
        result.repair_log.append(f"header_rows_skipped: {header_rows_skipped}")

    result.repair_log.append(
        f"parsed: {len(result.transactions)} transactions, "
        f"opening={result.opening_balance}, closing={result.closing_balance}"
    )

    wb.close()
    return result
