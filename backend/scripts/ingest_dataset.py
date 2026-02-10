#!/usr/bin/env python3
"""
Aegis - Dataset ingestion script.

Uses the same pipeline as the API to ingest local dataset folders.
"""

from __future__ import annotations

import argparse
import asyncio
import random
import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from sqlmodel import Session

from app.core.config import get_settings
from app.db.session import engine, init_db
from app.db.models import Document
from app.services.backboard_client import BackboardClient
from app.services.entity_resolution import resolve_entities
from app.services.file_preprocess import normalize_input
from app.services.layout import detect_layout_flags
from app.services.quality import score_image_quality
from app.services.storage import save_file
from app.services.validation import run_validations
from app.services.knowledge_graph import build_graph_from_document, get_knowledge_store, KGNode


VALID_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".xlsx"}
LABEL_MAP = {
    "Bank Statement": "bank_statement",
    "Check": "check",
    "ITR_Form 16": "form_16",
    "Salary Slip": "payslip",
    "Utility": "utility_bill",
}


def iter_dataset_files(
    dataset_root: Path,
    max_per_class: int | None,
    shuffle: bool,
    label_override: Optional[str] = None,
) -> Iterable[tuple[Path, str]]:
    folders = [path for path in dataset_root.iterdir() if path.is_dir()]
    if not folders:
        label = label_override or LABEL_MAP.get(
            dataset_root.name, dataset_root.name.lower().replace(" ", "_")
        )
        files = [
            path
            for path in dataset_root.iterdir()
            if path.is_file() and path.suffix.lower() in VALID_EXTENSIONS
        ]
        if shuffle:
            random.shuffle(files)
        if max_per_class:
            files = files[:max_per_class]
        for file_path in files:
            yield file_path, label
        return

    for folder in sorted(folders, key=lambda p: p.name.lower()):
        label = label_override or LABEL_MAP.get(folder.name, folder.name.lower().replace(" ", "_"))
        files = [
            path
            for path in folder.iterdir()
            if path.is_file() and path.suffix.lower() in VALID_EXTENSIONS
        ]
        if shuffle:
            random.shuffle(files)
        if max_per_class:
            files = files[:max_per_class]
        for file_path in files:
            yield file_path, label


def _read_xlsx_text(file_path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(value.strip() for value in row_values):
                lines.append("\t".join(row_values))
    return "\n".join(lines)


def _parse_amount(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    neg = False
    if text.startswith("(") and text.endswith(")"):
        neg = True
        text = text[1:-1]
    text = text.replace(",", "").replace("$", "").replace("₹", "").replace("€", "").strip()
    try:
        amount = float(text)
    except ValueError:
        return None
    return -amount if neg else amount


def _parse_date_value(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    try:
        from dateutil.parser import parse

        return parse(text, dayfirst=False, fuzzy=True).date().isoformat()
    except Exception:
        return None


def _detect_currency(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value)
    if "$" in text:
        return "USD"
    if "₹" in text:
        return "INR"
    if "€" in text:
        return "EUR"
    return None


def _find_header_row(rows: list[list[object]]) -> tuple[int | None, dict[str, int]]:
    header_map: dict[str, int] = {}
    best_row = None
    best_score = 0
    keywords = {
        "date": ["date", "txn date", "posting date", "value date"],
        "description": ["description", "narration", "details", "particulars", "remarks"],
        "debit": ["debit", "withdrawal", "withdraw", "dr", "paid out"],
        "credit": ["credit", "deposit", "cr", "paid in"],
        "amount": ["amount", "amt"],
        "balance": ["balance", "closing balance", "running balance"],
    }
    for idx, row in enumerate(rows[:30]):
        row_map: dict[str, int] = {}
        score = 0
        for col_idx, cell in enumerate(row):
            cell_text = str(cell).strip().lower() if cell is not None else ""
            for key, options in keywords.items():
                if any(opt in cell_text for opt in options):
                    row_map[key] = col_idx
        if "date" in row_map:
            score += 2
        if "amount" in row_map or ("debit" in row_map and "credit" in row_map):
            score += 2
        if "description" in row_map:
            score += 1
        if score > best_score:
            best_score = score
            best_row = idx
            header_map = row_map
    if best_score < 3:
        return None, {}
    return best_row, header_map


def _extract_transactions_from_xlsx(file_path: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    transactions: list[dict] = []
    opening_balance = None
    closing_balance = None
    currency = None
    account_number = None

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_idx, header_map = _find_header_row(rows)
        if header_idx is None:
            continue

        for row in rows[header_idx + 1 :]:
            date_value = _parse_date_value(row[header_map["date"]]) if "date" in header_map else None
            description = (
                str(row[header_map["description"]]).strip()
                if "description" in header_map and row[header_map["description"]] is not None
                else ""
            )

            debit = _parse_amount(row[header_map["debit"]]) if "debit" in header_map else None
            credit = _parse_amount(row[header_map["credit"]]) if "credit" in header_map else None
            amount = _parse_amount(row[header_map["amount"]]) if "amount" in header_map else None
            balance = _parse_amount(row[header_map["balance"]]) if "balance" in header_map else None

            if amount is None and (debit is not None or credit is not None):
                amount = (credit or 0.0) - (debit or 0.0)

            if not date_value and not description and amount is None:
                continue

            detected_currency = _detect_currency(row[header_map["amount"]]) if "amount" in header_map else None
            if detected_currency and not currency:
                currency = detected_currency

            if balance is not None:
                if opening_balance is None:
                    opening_balance = balance
                closing_balance = balance

            transactions.append(
                {
                    "date": date_value,
                    "amount": amount,
                    "currency": detected_currency or currency,
                    "description": description,
                    "balance": balance,
                }
            )

        if transactions:
            break

    if account_number is None:
        account_pattern = re.compile(r"\b\d{8,18}\b")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    match = account_pattern.search(str(cell)) if cell is not None else None
                    if match:
                        account_number = match.group(0)
                        break
                if account_number:
                    break
            if account_number:
                break

    return {
        "transactions": transactions,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "currency": currency,
        "account_number": account_number,
    }


def _score_bank_statement_text(text: str) -> int:
    if not text:
        return 0
    lowered = text.lower()
    keywords = [
        "opening balance",
        "closing balance",
        "statement period",
        "account number",
        "debit",
        "credit",
        "transaction",
        "balance",
        "date",
    ]
    return sum(1 for kw in keywords if kw in lowered)


async def ingest_file(
    file_path: Path,
    label: str,
    session: Session,
    client: BackboardClient,
    doc_hint: Optional[str] = None,
) -> dict:
    settings = get_settings()
    content = file_path.read_bytes()
    if not content:
        return {"filename": file_path.name, "status": "failed", "error": "Empty file"}

    if len(content) > settings.max_upload_mb * 1024 * 1024:
        return {"filename": file_path.name, "status": "failed", "error": "File too large"}

    ext = file_path.suffix.lower()
    quality_metrics = score_image_quality(content) if ext != ".xlsx" else {}
    local_layout = detect_layout_flags(content) if ext != ".xlsx" else {}

    try:
        if ext == ".xlsx":
            try:
                text = _read_xlsx_text(file_path)
            except Exception as exc:
                return {"filename": file_path.name, "status": "failed", "error": f"Excel read failed: {exc}"}
            analysis = await client.analyze_text(text, doc_hint=doc_hint)
            local_fields = _extract_transactions_from_xlsx(file_path)
            normalized = None
        else:
            normalized = normalize_input(file_path.name, content)
            analysis = await client.analyze_document(
                normalized.normalized_bytes,
                normalized.normalized_name,
                mime_type=normalized.normalized_mime,
                fallback_bytes=normalized.original_bytes if normalized.converted else None,
                fallback_filename=normalized.original_name if normalized.converted else None,
                fallback_mime=normalized.original_mime if normalized.converted else None,
                doc_hint=doc_hint,
            )
    except Exception as exc:
        return {"filename": file_path.name, "status": "failed", "error": str(exc)}

    classification = analysis.get("classification", {})
    extracted_fields = analysis.get("extracted_fields", {})
    if ext == ".xlsx":
        local_transactions = local_fields.get("transactions") or []
        if not extracted_fields.get("transactions") and local_transactions:
            extracted_fields = {**extracted_fields, "transactions": local_transactions}
        for key in ("opening_balance", "closing_balance", "account_number"):
            if extracted_fields.get(key) in (None, "") and local_fields.get(key) is not None:
                extracted_fields = {**extracted_fields, key: local_fields[key]}
        if local_fields.get("currency") and not extracted_fields.get("currency"):
            extracted_fields = {**extracted_fields, "currency": local_fields["currency"]}
    remote_layout = analysis.get("layout", {})
    layout_flags = {**local_layout, **remote_layout}
    backboard_thread_id = analysis.get("document_id")
    parse_error = analysis.get("parse_error")

    doc_type = classification.get("type", "unknown")
    if doc_type in (None, "unknown") and label == "bank_statement" and ext == ".xlsx":
        if _score_bank_statement_text(text) >= 3:
            doc_type = "bank_statement"
            classification = {**classification, "type": doc_type}

    errors, warnings, consistency = run_validations(
        doc_type,
        extracted_fields,
        session,
    )
    if parse_error:
        warnings.append({"field": "backboard", "message": parse_error})

    confidence = classification.get("confidence") or 0.0
    image_quality = classification.get("image_quality_score") or quality_metrics.get("score")
    status = "processed"
    if errors or warnings:
        status = "review"
    if confidence < settings.review_confidence_threshold:
        status = "review"
    if image_quality is not None and image_quality < settings.review_quality_threshold:
        status = "review"
    if doc_type in (None, "unknown"):
        status = "review"

    merged_consistency = dict(consistency)
    merged_consistency["dataset_label"] = label
    merged_consistency["source_path"] = str(file_path)

    doc = Document(
        filename=normalized.normalized_name if normalized else file_path.name,
        backboard_thread_id=backboard_thread_id,
        doc_type=doc_type,
        confidence=confidence,
        language=classification.get("language"),
        image_quality=image_quality,
        status=status,
        layout_flags=layout_flags,
        quality_metrics=quality_metrics,
        extracted_fields=extracted_fields,
        validation_errors=errors,
        validation_warnings=warnings,
        consistency=merged_consistency,
    )
    doc.file_path = save_file(
        doc.id,
        doc.filename,
        normalized.normalized_bytes if normalized else content,
    )
    session.add(doc)
    session.commit()
    session.refresh(doc)

    entities = resolve_entities(session, doc.id, extracted_fields)
    entity_nodes = [
        KGNode(
            id=entity.id,
            type=entity.entity_type,
            properties={"canonical_value": entity.canonical_value},
        )
        for entity in entities
    ]
    graph = build_graph_from_document(doc.id, doc.doc_type, extracted_fields, entity_nodes)
    get_knowledge_store().upsert_document_graph(graph)

    return {
        "document_id": doc.id,
        "filename": doc.filename,
        "doc_type": doc.doc_type,
        "confidence": doc.confidence,
        "status": status,
        "layout": layout_flags,
        "quality_metrics": quality_metrics,
        "error": parse_error,
        "dataset_label": label,
    }


async def ingest_dataset(
    dataset_root: Path,
    max_per_class: int | None,
    limit: int | None,
    shuffle: bool,
    dry_run: bool,
    label_override: Optional[str] = None,
) -> list[dict]:
    settings = get_settings()
    if not settings.backboard_api_key:
        raise RuntimeError("BACKBOARD_API_KEY is not configured.")

    init_db()
    results: list[dict] = []
    client = BackboardClient()

    with Session(engine) as session:
        for index, (file_path, label) in enumerate(
            iter_dataset_files(dataset_root, max_per_class, shuffle, label_override)
        ):
            if limit and index >= limit:
                break
            if dry_run:
                results.append(
                    {
                        "filename": file_path.name,
                        "status": "skipped",
                        "dataset_label": label,
                    }
                )
                continue
            doc_hint = None
            if (label_override or label) == "bank_statement":
                doc_hint = (
                    "This document is a bank statement. Extract account holder, "
                    "account number, statement period, opening/closing balances, "
                    "and transaction lines."
                )
            result = await ingest_file(file_path, label, session, client, doc_hint=doc_hint)
            results.append(result)

    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest dataset folder into Aegis")
    parser.add_argument(
        "--dataset-path",
        type=str,
        default=None,
        help="Path to dataset root (defaults to DATASET_PATH env or settings).",
    )
    parser.add_argument(
        "--max-per-class",
        type=int,
        default=10,
        help="Max files to ingest per class folder.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional global limit across all files.",
    )
    parser.add_argument(
        "--shuffle",
        action="store_true",
        help="Shuffle files within each class.",
    )
    parser.add_argument(
        "--label",
        type=str,
        default=None,
        help="Optional label override for all files (e.g., bank_statement).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List files that would be ingested without calling the API.",
    )
    return parser.parse_args()


def resolve_dataset_path(args: argparse.Namespace) -> Path:
    settings = get_settings()
    if args.dataset_path:
        return Path(args.dataset_path).expanduser()
    if settings.dataset_path:
        return Path(settings.dataset_path).expanduser()
    raise RuntimeError("Dataset path not provided. Use --dataset-path or DATASET_PATH.")


def main() -> None:
    args = parse_args()
    dataset_root = resolve_dataset_path(args)
    if not dataset_root.exists():
        raise RuntimeError(f"Dataset path not found: {dataset_root}")

    results = asyncio.run(
        ingest_dataset(
            dataset_root=dataset_root,
            max_per_class=args.max_per_class,
            limit=args.limit,
            shuffle=args.shuffle,
            dry_run=args.dry_run,
            label_override=args.label,
        )
    )

    total = len(results)
    failed = sum(1 for item in results if item.get("status") == "failed")
    reviewed = sum(1 for item in results if item.get("status") == "review")
    processed = sum(1 for item in results if item.get("status") == "processed")

    print("\n📦 Dataset ingestion complete")
    print(f"Total: {total} | Processed: {processed} | Review: {reviewed} | Failed: {failed}")
    if failed:
        print("\n❌ Failed files:")
        for item in results:
            if item.get("status") == "failed":
                print(f"- {item.get('filename')}: {item.get('error')}")


if __name__ == "__main__":
    main()
